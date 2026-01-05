from openpyxl import load_workbook

try:
    # Load the workbook
    wb = load_workbook('EOR Switch Details - Current.xlsx')
    
    # Get the active sheet
    sheet = wb.active
    
    # Print sheet information
    print(f"Sheet name: {sheet.title}")
    print(f"Columns: {[cell.value for cell in sheet[1]]}")
    
    # Print first 5 rows of data
    print("\nFirst 5 rows of data:")
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        print(row)
        if i >= 5:
            break
            
except Exception as e:
    print(f"Error: {str(e)}")
